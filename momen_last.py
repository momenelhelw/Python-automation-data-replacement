#!/usr/bin/env python
# coding: utf-8

# In[1]:


from openpyxl import load_workbook
from openpyxl import Workbook

wb = Workbook()
book = load_workbook('Copy of Generationfile(114).xlsx')                    #Read the excel file name is book
print(book.sheetnames)
sheet = book.active

row = sheet.max_row
column = sheet.max_column

print("Total Rows:", row)
print("Total Columns:", column)


# In[2]:


list_file_name = []
for j in range(2, sheet.max_row + 1):
    cell1 = sheet.cell(row=j, column=2)
    outputFile = open('{}.xml'.format(cell1.value), 'w')
    with open("CanComm_RX_[ESP_HL_Radgeschw_02].xml", 'r') as file:         #READ THE XML FILE
        # string = file.read().replace('ESP_HL_Radgeschw_02', 'BMS_IstModus')
        string = file.read()
    with open('{}.xml'.format(cell1.value), 'w') as a:                      #READ AND WRITE THE NAMES OF NEW XMLS
        a.write(string)
        list_file_name.append(cell1.value)
        # outputFile.write(cell1.value + '\n')

outputFile.close()
print("Total files:",len(list_file_name),"\n",list_file_name)


# In[3]:


for y in range(3, sheet.max_column + 1, 2):
    for j in range(2, sheet.max_row + 1):
        cell2 = sheet.cell(row=j, column=y)
        cell3 = sheet.cell(row=j, column=y + 1)
        str_cell2=str(cell2.value)
        str_cell3=str(cell3.value)

        search_text = str_cell2
        replace_text = str_cell3

        with open(r'{}.xml'.format(list_file_name[j - 2]), 'r') as file:     #READ AND EDIT DATA IN XML FILES
            data = file.read()
            data = data.replace(search_text, replace_text)

        with open(r'{}.xml'.format(list_file_name[j - 2]), 'w') as file:
            file.write(data)


# In[ ]:




